"""
accounts/exports.py
Gestion centralisée de tous les exports (PDF / Excel) pour l'application accounts.
"""
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect


# ──────────────────────────────────────────────────────────────────────────────
# Utilitaires communs
# ──────────────────────────────────────────────────────────────────────────────

def _get_reportlab():
    """Import lazy de ReportLab pour éviter les erreurs si non installé."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        return {
            'A4': A4, 'landscape': landscape, 'colors': colors,
            'inch': inch, 'cm': cm,
            'SimpleDocTemplate': SimpleDocTemplate, 'Table': Table,
            'TableStyle': TableStyle, 'Paragraph': Paragraph,
            'Spacer': Spacer, 'HRFlowable': HRFlowable,
            'getSampleStyleSheet': getSampleStyleSheet,
            'ParagraphStyle': ParagraphStyle,
            'TA_CENTER': TA_CENTER, 'TA_LEFT': TA_LEFT, 'TA_RIGHT': TA_RIGHT,
        }
    except ImportError:
        raise ImportError("ReportLab n'est pas installé. Exécutez : pip install reportlab")


def _get_openpyxl():
    """Import lazy de openpyxl pour éviter les erreurs si non installé."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        return {
            'Workbook': Workbook, 'Font': Font, 'PatternFill': PatternFill,
            'Alignment': Alignment, 'Border': Border, 'Side': Side,
            'get_column_letter': get_column_letter,
        }
    except ImportError:
        raise ImportError("openpyxl n'est pas installé. Exécutez : pip install openpyxl")


def _pdf_response(buffer, filename):
    """Retourne une HttpResponse PDF à télécharger."""
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _excel_response(buffer, filename):
    """Retourne une HttpResponse Excel à télécharger."""
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _build_pdf_doc(buffer, title_text, rl):
    """Crée un document PDF avec un style de titre commun."""
    doc = rl['SimpleDocTemplate'](
        buffer,
        pagesize=rl['A4'],
        rightMargin=1.5 * rl['cm'],
        leftMargin=1.5 * rl['cm'],
        topMargin=2 * rl['cm'],
        bottomMargin=2 * rl['cm'],
    )
    styles = rl['getSampleStyleSheet']()

    title_style = rl['ParagraphStyle'](
        'LonabTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=rl['colors'].HexColor('#1a3a2a'),
        spaceAfter=6,
        alignment=rl['TA_CENTER'],
        fontName='Helvetica-Bold',
    )
    subtitle_style = rl['ParagraphStyle'](
        'LonabSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=rl['colors'].HexColor('#6b7280'),
        spaceAfter=20,
        alignment=rl['TA_CENTER'],
    )

    elements = [
        rl['Paragraph']('LONAB — MUTRALO', title_style),
        rl['Paragraph'](title_text, subtitle_style),
        rl['Paragraph'](f'Généré le {timezone.now().strftime("%d/%m/%Y à %H:%M")}', subtitle_style),
        rl['HRFlowable'](width='100%', thickness=2, color=rl['colors'].HexColor('#1e5c3a'), spaceAfter=16),
    ]
    return doc, elements


def _table_style(rl, header_rows=1):
    """Style commun pour les tableaux PDF."""
    return rl['TableStyle']([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, header_rows - 1), rl['colors'].HexColor('#1e5c3a')),
        ('TEXTCOLOR', (0, 0), (-1, header_rows - 1), rl['colors'].white),
        ('FONTNAME', (0, 0), (-1, header_rows - 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, header_rows - 1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, header_rows - 1), 10),
        ('TOPPADDING', (0, 0), (-1, header_rows - 1), 10),
        # Corps
        ('FONTNAME', (0, header_rows), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, header_rows), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, header_rows), (-1, -1), [
            rl['colors'].white, rl['colors'].HexColor('#f2faf5')
        ]),
        # Grille
        ('GRID', (0, 0), (-1, -1), 0.5, rl['colors'].HexColor('#e5e7eb')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, rl['colors'].HexColor('#1e5c3a')),
        # Padding
        ('TOPPADDING', (0, header_rows), (-1, -1), 7),
        ('BOTTOMPADDING', (0, header_rows), (-1, -1), 7),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        # Alignement
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])


def _excel_header_style(xl):
    """Style d'en-tête commun pour les fichiers Excel."""
    return {
        'font': xl['Font'](bold=True, color='FFFFFF', size=10),
        'fill': xl['PatternFill'](start_color='1e5c3a', end_color='1e5c3a', fill_type='solid'),
        'alignment': xl['Alignment'](horizontal='center', vertical='center', wrap_text=True),
    }


def _excel_apply_header(ws, headers, xl):
    """Applique les en-têtes Excel avec le style LONAB."""
    style = _excel_header_style(xl)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']


def _excel_add_alt_rows(ws, start_row, xl):
    """Ajoute des lignes alternées pour une meilleure lisibilité."""
    alt_fill = xl['PatternFill'](start_color='f2faf5', end_color='f2faf5', fill_type='solid')
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        for cell in row:
            if i % 2 == 1:
                cell.fill = alt_fill
            cell.alignment = xl['Alignment'](vertical='center', wrap_text=False)


def _excel_set_widths(ws, widths):
    """Définit la largeur des colonnes Excel."""
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Hauteur de la ligne d'en-tête
    ws.row_dimensions[1].height = 24


# ──────────────────────────────────────────────────────────────────────────────
# Exports Utilisateurs
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def export_users_pdf(request):
    """Exporter la liste des utilisateurs en PDF."""
    if not request.user.est_admin:
        return redirect('accounts:users_list')

    from .models import Utilisateur
    rl = _get_reportlab()
    buffer = BytesIO()
    doc, elements = _build_pdf_doc(buffer, 'Liste des Utilisateurs', rl)

    # Filtres appliqués (mêmes que la vue)
    qs = Utilisateur.objects.all().select_related('direction', 'agence')
    type_filter = request.GET.get('type')
    if type_filter:
        qs = qs.filter(type_utilisateur=type_filter)
    direction_filter = request.GET.get('direction')
    if direction_filter:
        qs = qs.filter(direction_id=direction_filter)
    agence_filter = request.GET.get('agence')
    if agence_filter:
        qs = qs.filter(agence_id=agence_filter)
    statut_filter = request.GET.get('statut')
    if statut_filter == 'actif':
        qs = qs.filter(est_actif=True)
    elif statut_filter == 'inactif':
        qs = qs.filter(est_actif=False)
    from django.db.models import Q
    search = request.GET.get('search')
    if search:
        qs = qs.filter(Q(prenom__icontains=search)|Q(nom__icontains=search)|Q(email__icontains=search)|Q(matricule__icontains=search))

    # Compteur de résumé
    elements.append(rl['Paragraph'](f'{qs.count()} utilisateur(s) exporté(s)', rl['getSampleStyleSheet']()['Normal']))
    elements.append(rl['Spacer'](1, 12))

    # Tableau
    headers = ['#', 'Nom complet', 'Email', 'Type', 'Matricule', 'Direction', 'Agence', 'Statut']
    data = [headers]
    for idx, user in enumerate(qs.order_by('-date_inscription'), 1):
        data.append([
            str(idx),
            user.get_full_name(),
            user.email,
            user.get_type_utilisateur_display(),
            user.matricule or '—',
            user.direction.nom if user.direction else '—',
            user.agence.nom if user.agence else '—',
            'Actif' if user.est_actif else 'Inactif',
        ])

    col_widths = [0.4 * rl['inch'], 1.5 * rl['inch'], 2 * rl['inch'], 1.2 * rl['inch'],
                  0.9 * rl['inch'], 1.2 * rl['inch'], 1.2 * rl['inch'], 0.7 * rl['inch']]
    table = rl['Table'](data, colWidths=col_widths, repeatRows=1)
    table.setStyle(_table_style(rl))
    elements.append(table)
    doc.build(elements)

    filename = f"utilisateurs_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    return _pdf_response(buffer, filename)


@login_required
def export_users_excel(request):
    """Exporter la liste des utilisateurs en Excel."""
    if not request.user.est_admin:
        return redirect('accounts:users_list')

    from .models import Utilisateur
    xl = _get_openpyxl()
    wb = xl['Workbook']()
    ws = wb.active
    ws.title = 'Utilisateurs'

    # Filtres (mêmes que la vue)
    qs = Utilisateur.objects.all().select_related('direction', 'agence')
    type_filter = request.GET.get('type')
    if type_filter:
        qs = qs.filter(type_utilisateur=type_filter)
    direction_filter = request.GET.get('direction')
    if direction_filter:
        qs = qs.filter(direction_id=direction_filter)
    agence_filter = request.GET.get('agence')
    if agence_filter:
        qs = qs.filter(agence_id=agence_filter)
    statut_filter = request.GET.get('statut')
    if statut_filter == 'actif':
        qs = qs.filter(est_actif=True)
    elif statut_filter == 'inactif':
        qs = qs.filter(est_actif=False)
    from django.db.models import Q
    search = request.GET.get('search')
    if search:
        qs = qs.filter(Q(prenom__icontains=search)|Q(nom__icontains=search)|Q(email__icontains=search)|Q(matricule__icontains=search))

    headers = ['#', 'Prénom', 'Nom', 'Email', 'Téléphone', 'Type', 'Matricule',
               'Département', 'Poste', 'Direction', 'Agence', 'Statut', 'Inscrit le']
    _excel_apply_header(ws, headers, xl)

    for idx, user in enumerate(qs.order_by('-date_inscription'), 1):
        ws.append([
            idx, user.prenom, user.nom, user.email,
            user.telephone or '—', user.get_type_utilisateur_display(),
            user.matricule or '—', user.departement or '—', user.poste or '—',
            user.direction.nom if user.direction else '—',
            user.agence.nom if user.agence else '—',
            'Actif' if user.est_actif else 'Inactif',
            user.date_inscription.strftime('%d/%m/%Y') if user.date_inscription else '—',
        ])

    _excel_add_alt_rows(ws, 2, xl)
    _excel_set_widths(ws, [5, 16, 16, 28, 16, 22, 14, 18, 18, 22, 22, 10, 14])

    buffer = BytesIO()
    wb.save(buffer)
    filename = f"utilisateurs_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return _excel_response(buffer, filename)


# ──────────────────────────────────────────────────────────────────────────────
# Exports Directions
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def export_directions_pdf(request):
    """Exporter la liste des directions en PDF."""
    if not request.user.est_admin:
        return redirect('accounts:directions_list')

    from .models import Direction
    from django.db.models import Count, Q
    rl = _get_reportlab()
    buffer = BytesIO()
    doc, elements = _build_pdf_doc(buffer, 'Liste des Directions', rl)

    directions = Direction.objects.all().annotate(
        nombre_employes=Count('employes', filter=Q(employes__type_utilisateur='CLIENT', employes__est_actif=True)),
        nombre_agences=Count('agences', filter=Q(agences__est_active=True)),
    ).order_by('nom')

    search = request.GET.get('search')
    if search:
        directions = directions.filter(Q(nom__icontains=search)|Q(code__icontains=search))
    statut = request.GET.get('statut')
    if statut == 'actif':
        directions = directions.filter(est_active=True)
    elif statut == 'inactif':
        directions = directions.filter(est_active=False)

    elements.append(rl['Paragraph'](f'{directions.count()} direction(s) exportée(s)', rl['getSampleStyleSheet']()['Normal']))
    elements.append(rl['Spacer'](1, 12))

    headers = ['#', 'Nom', 'Code', 'Directeur', 'Employés actifs', 'Agences', 'Statut']
    data = [headers]
    for idx, d in enumerate(directions, 1):
        data.append([
            str(idx), d.nom, d.code,
            d.directeur.get_full_name() if d.directeur else '—',
            str(d.nombre_employes), str(d.nombre_agences),
            'Active' if d.est_active else 'Inactive',
        ])

    col_widths = [0.4 * rl['inch'], 2.2 * rl['inch'], 0.9 * rl['inch'], 1.5 * rl['inch'],
                  1.1 * rl['inch'], 0.8 * rl['inch'], 0.8 * rl['inch']]
    table = rl['Table'](data, colWidths=col_widths, repeatRows=1)
    table.setStyle(_table_style(rl))
    elements.append(table)
    doc.build(elements)

    filename = f"directions_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    return _pdf_response(buffer, filename)


@login_required
def export_directions_excel(request):
    """Exporter la liste des directions en Excel."""
    if not request.user.est_admin:
        return redirect('accounts:directions_list')

    from .models import Direction
    from django.db.models import Count, Q
    xl = _get_openpyxl()
    wb = xl['Workbook']()
    ws = wb.active
    ws.title = 'Directions'

    directions = Direction.objects.all().annotate(
        nombre_employes=Count('employes', filter=Q(employes__type_utilisateur='CLIENT', employes__est_actif=True)),
        nombre_agences=Count('agences', filter=Q(agences__est_active=True)),
    ).order_by('nom')

    headers = ['#', 'Nom', 'Code', 'Description', 'Directeur', 'Téléphone', 'Email',
               'Employés actifs', 'Agences liées', 'Statut', 'Créée le']
    _excel_apply_header(ws, headers, xl)

    for idx, d in enumerate(directions, 1):
        ws.append([
            idx, d.nom, d.code, d.description or '—',
            d.directeur.get_full_name() if d.directeur else '—',
            d.telephone or '—', d.email or '—',
            d.nombre_employes, d.nombre_agences,
            'Active' if d.est_active else 'Inactive',
            d.date_creation.strftime('%d/%m/%Y') if d.date_creation else '—',
        ])

    _excel_add_alt_rows(ws, 2, xl)
    _excel_set_widths(ws, [5, 28, 12, 35, 22, 16, 24, 14, 14, 10, 14])

    buffer = BytesIO()
    wb.save(buffer)
    filename = f"directions_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return _excel_response(buffer, filename)


# ──────────────────────────────────────────────────────────────────────────────
# Exports Agences
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def export_agencies_pdf(request):
    """Exporter la liste des agences en PDF."""
    if not request.user.est_admin:
        return redirect('accounts:agences_list')

    from .models import Agence
    from django.db.models import Count, Q
    rl = _get_reportlab()
    buffer = BytesIO()
    doc, elements = _build_pdf_doc(buffer, 'Liste des Agences', rl)

    agences = Agence.objects.all().select_related('direction', 'responsable').annotate(
        nombre_employes=Count('employes', filter=Q(employes__type_utilisateur='CLIENT', employes__est_actif=True))
    ).order_by('nom')

    search = request.GET.get('search')
    if search:
        agences = agences.filter(Q(nom__icontains=search)|Q(code__icontains=search)|Q(ville__icontains=search))
    type_f = request.GET.get('type')
    if type_f:
        agences = agences.filter(type_agence=type_f)
    statut = request.GET.get('statut')
    if statut == 'actif':
        agences = agences.filter(est_active=True)
    elif statut == 'inactif':
        agences = agences.filter(est_active=False)

    elements.append(rl['Paragraph'](f'{agences.count()} agence(s) exportée(s)', rl['getSampleStyleSheet']()['Normal']))
    elements.append(rl['Spacer'](1, 12))

    headers = ['#', 'Nom', 'Code', 'Type', 'Ville', 'Direction', 'Responsable', 'Employés', 'Statut']
    data = [headers]
    for idx, a in enumerate(agences, 1):
        data.append([
            str(idx), a.nom, a.code, a.get_type_agence_display(),
            a.ville, a.direction.nom if a.direction else '—',
            a.responsable.get_full_name() if a.responsable else '—',
            str(a.nombre_employes), 'Active' if a.est_active else 'Inactive',
        ])

    col_widths = [0.4 * rl['inch'], 1.8 * rl['inch'], 0.8 * rl['inch'], 1 * rl['inch'],
                  0.9 * rl['inch'], 1.2 * rl['inch'], 1.2 * rl['inch'], 0.7 * rl['inch'], 0.7 * rl['inch']]
    table = rl['Table'](data, colWidths=col_widths, repeatRows=1)
    table.setStyle(_table_style(rl))
    elements.append(table)
    doc.build(elements)

    filename = f"agences_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    return _pdf_response(buffer, filename)


@login_required
def export_agencies_excel(request):
    """Exporter la liste des agences en Excel."""
    if not request.user.est_admin:
        return redirect('accounts:agences_list')

    from .models import Agence
    from django.db.models import Count, Q
    xl = _get_openpyxl()
    wb = xl['Workbook']()
    ws = wb.active
    ws.title = 'Agences'

    agences = Agence.objects.all().select_related('direction', 'responsable').annotate(
        nombre_employes=Count('employes', filter=Q(employes__type_utilisateur='CLIENT', employes__est_actif=True))
    ).order_by('nom')

    headers = ['#', 'Nom', 'Code', 'Type', 'Adresse', 'Ville', 'Région',
               'Téléphone', 'Email', 'Direction', 'Responsable', 'Employés', 'Statut', 'Ouverte le']
    _excel_apply_header(ws, headers, xl)

    for idx, a in enumerate(agences, 1):
        ws.append([
            idx, a.nom, a.code, a.get_type_agence_display(),
            a.adresse or '—', a.ville, a.region or '—',
            a.telephone or '—', a.email or '—',
            a.direction.nom if a.direction else '—',
            a.responsable.get_full_name() if a.responsable else '—',
            a.nombre_employes,
            'Active' if a.est_active else 'Inactive',
            a.date_ouverture.strftime('%d/%m/%Y') if a.date_ouverture else '—',
        ])

    _excel_add_alt_rows(ws, 2, xl)
    _excel_set_widths(ws, [5, 22, 10, 14, 28, 16, 16, 16, 24, 22, 22, 10, 10, 14])

    buffer = BytesIO()
    wb.save(buffer)
    filename = f"agences_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return _excel_response(buffer, filename)

