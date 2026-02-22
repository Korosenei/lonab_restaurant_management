"""
Vues pour l'application comptes
"""
import secrets
import string

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login as auth_login, logout as auth_logout, authenticate
from django.contrib.auth import authenticate as django_authenticate
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.backends import ModelBackend
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime, timedelta
from io import BytesIO
import json

from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.tokens import RefreshToken

from .models import Utilisateur, Direction, Agence, ProfilUtilisateur
from .serializers import (
    UtilisateurSerializer, UtilisateurCreateSerializer, LoginSerializer,
    DirectionSerializer, AgenceSerializer, ProfilUtilisateurSerializer,
    ChangerMotDePasseSerializer, ReinitialisationMotDePasseSerializer,
    UtilisateurMiseAJourSerializer, ProfilUtilisateurMiseAJourSerializer
)


# ============================================
# Authentification email OU username
# ============================================
class EmailOrUsernameBackend(ModelBackend):
    def authenticate(self, request, username=None, password=None, **kwargs):
        if not username or not password:
            return None

        # Chercher par email
        user = None
        try:
            user = Utilisateur.objects.get(email=username)
        except Utilisateur.DoesNotExist:
            pass

        # Sinon par nom_utilisateur (matricule)
        if user is None:
            try:
                user = Utilisateur.objects.get(nom_utilisateur=username)
            except Utilisateur.DoesNotExist:
                return None

        # Vérifier le mot de passe (Django utilise PBKDF2 par défaut — set_password crypte, check_password décrypte)
        if user.check_password(password) and self.user_can_authenticate(user):
            return user
        return None

# ============================================
# Vues d'authentification
# ============================================
@ensure_csrf_cookie
@never_cache
@require_http_methods(["GET", "POST"])
def login_view(request):
    """Connexion par email OU par nom_utilisateur (matricule)"""
    if request.user.is_authenticated:
        return redirect('accounts:dashboard')

    if request.method == 'POST':
        identifiant = request.POST.get('email', '').strip()   # champ email accepte les deux
        mot_de_passe = request.POST.get('password', '')
        se_souvenir = request.POST.get('remember_me', False)

        if not identifiant or not mot_de_passe:
            messages.error(request, 'Veuillez remplir tous les champs.')
            return render(request, 'auth/login.html')

        # Django authenticate utilise le backend enregistré dans AUTHENTICATION_BACKENDS
        utilisateur = django_authenticate(request, username=identifiant, password=mot_de_passe)

        if utilisateur is not None:
            if utilisateur.is_active:
                auth_login(request, utilisateur)
                utilisateur.derniere_connexion = timezone.now()
                utilisateur.save(update_fields=['derniere_connexion'])

                if not se_souvenir:
                    request.session.set_expiry(0)
                else:
                    request.session.set_expiry(1209600)  # 14 jours

                messages.success(request, f'Bienvenue {utilisateur.get_full_name()} !')

                if utilisateur.est_super_admin:
                    return redirect('accounts:admin_dashboard')
                elif utilisateur.est_caissier:
                    return redirect('accounts:cashier_dashboard')
                elif utilisateur.est_gestionnaire_restaurant:
                    return redirect('accounts:restaurant_dashboard')
                else:
                    return redirect('accounts:client_dashboard')
            else:
                messages.error(request, 'Ce compte est désactivé.')
        else:
            messages.error(request, 'Identifiant ou mot de passe incorrect.')

    return render(request, 'auth/login.html')

@require_http_methods(["GET", "POST"])
def logout_view(request):
    """Vue de déconnexion"""
    auth_logout(request)
    messages.success(request, 'Vous avez été déconnecté avec succès.')
    return redirect('accounts:login')

@ensure_csrf_cookie
@require_http_methods(["GET", "POST"])
def password_reset_view(request):
    """Vue de demande de réinitialisation de mot de passe"""
    if request.method == 'POST':
        serializer = ReinitialisationMotDePasseSerializer(data=request.POST)

        if serializer.is_valid():
            email = serializer.validated_data['email']
            # TODO: Envoyer l'email de réinitialisation
            messages.success(request, 'Un email de réinitialisation a été envoyé à votre adresse.')
            return redirect('accounts:login')
        else:
            for error in serializer.errors.values():
                messages.error(request, error[0])

    return render(request, 'auth/password_reset.html')


# ============================================
# Vues de tableau de bord
# ============================================
@login_required
def dashboard_redirection(request):
    """Rediriger vers le tableau de bord approprié selon le type d'utilisateur"""
    utilisateur = request.user

    if utilisateur.est_super_admin:
        return redirect('accounts:admin_dashboard')
    elif utilisateur.est_caissier:
        return redirect('accounts:caissier_dashboard')
    elif utilisateur.est_gestionnaire_restaurant:
        return redirect('accounts:restaurant_dashboard')
    else:  # CLIENT
        return redirect('accounts:client_dashboard')

@login_required
def dashboard_client(request):
    """Tableau de bord pour les clients (employés)"""
    if not request.user.est_client:
        messages.warning(request, 'Vous n\'avez pas accès à cette section.')
        return redirect('accounts:dashboard')

    utilisateur = request.user
    aujourdhui = timezone.now().date()

    # Importer les modèles ici pour éviter les imports circulaires
    try:
        from apps.tickets.models import Ticket, QRCode
        from apps.transactions.models import JournalConsommation
        from apps.restaurants.models import Reservation

        # Obtenir les statistiques des tickets
        tickets_disponibles = Ticket.objects.filter(
            proprietaire=utilisateur,
            statut='DISPONIBLE',
            valide_du__lte=aujourdhui,
            valide_jusqu_a__gte=aujourdhui
        )

        tickets_consommes = Ticket.objects.filter(
            proprietaire=utilisateur,
            statut='CONSOMME'
        )

        # Obtenir le code QR actif
        code_qr = QRCode.objects.filter(
            utilisateur=utilisateur,
            est_valide=True,
            est_utilise=False,
            expire_a__gt=timezone.now()
        ).first()

        # Obtenir l'activité récente
        activite_recente = JournalConsommation.objects.filter(
            client=utilisateur
        ).select_related('restaurant', 'menu_consomme').order_by('-date_consommation')[:5]

        # Obtenir les réservations actives
        reservations_actives = Reservation.objects.filter(
            client=utilisateur,
            statut__in=['EN_ATTENTE', 'CONFIRME'],
            date_reservation__gte=aujourdhui
        )

        contexte = {
            'nombre_tickets_disponibles': tickets_disponibles.count(),
            'nombre_tickets_consommes': tickets_consommes.count(),
            'valeur_totale': tickets_disponibles.count() * 2000,  # 2000 FCFA par ticket
            'nombre_reservations_actives': reservations_actives.count(),
            'code_qr': code_qr,
            'activite_recente': activite_recente,
        }
    except ImportError:
        # Si les apps n'existent pas encore
        contexte = {
            'nombre_tickets_disponibles': 0,
            'nombre_tickets_consommes': 0,
            'valeur_totale': 0,
            'nombre_reservations_actives': 0,
            'code_qr': None,
            'activite_recente': [],
        }

    contexte.update({
        'annee_courante': timezone.now().year,
        'notifications': [],
        'nombre_notifications_non_lues': 0,
    })

    return render(request, 'dashboards/client_dashboard.html', contexte)

@login_required
def dashboard_caissier(request):
    """Tableau de bord pour les caissiers"""
    if not request.user.est_caissier:
        messages.warning(request, 'Vous n\'avez pas accès à cette section.')
        return redirect('accounts:dashboard')

    try:
        from apps.transactions.models import TransactionTicket
        from apps.restaurants.models import Restaurant

        aujourdhui = timezone.now().date()
        debut_mois = aujourdhui.replace(day=1)

        # Obtenir les ventes du jour
        ventes_du_jour = TransactionTicket.objects.filter(
            caissier=request.user,
            date_transaction__date=aujourdhui,
            statut='TERMINE'
        )

        # Obtenir les ventes du mois
        ventes_du_mois = TransactionTicket.objects.filter(
            caissier=request.user,
            date_transaction__date__gte=debut_mois,
            statut='TERMINE'
        )

        # Obtenir les clients actifs
        clients_actifs = Utilisateur.objects.filter(
            type_utilisateur='CLIENT',
            est_actif=True,
            agence=request.user.agence
        ) if request.user.agence else Utilisateur.objects.filter(type_utilisateur='CLIENT', est_actif=True)

        # Obtenir les transactions récentes
        transactions_recentes = TransactionTicket.objects.filter(
            caissier=request.user
        ).select_related('client').order_by('-date_transaction')[:10]

        # Obtenir les restaurants actifs
        restaurants_actifs = Restaurant.objects.filter(
            statut='ACTIF'
        )[:5]

        contexte = {
            'nombre_ventes_jour': ventes_du_jour.count(),
            'montant_ventes_jour': sum(t.montant_total for t in ventes_du_jour),
            'nombre_ventes_mois': ventes_du_mois.count(),
            'montant_ventes_mois': sum(t.montant_total for t in ventes_du_mois),
            'tickets_vendus_mois': sum(t.nombre_tickets for t in ventes_du_mois),
            'nombre_clients_actifs': clients_actifs.count(),
            'transactions_recentes': transactions_recentes,
            'restaurants_actifs': restaurants_actifs,
        }
    except ImportError:
        contexte = {
            'nombre_ventes_jour': 0,
            'montant_ventes_jour': 0,
            'nombre_ventes_mois': 0,
            'montant_ventes_mois': 0,
            'tickets_vendus_mois': 0,
            'nombre_clients_actifs': 0,
            'transactions_recentes': [],
            'restaurants_actifs': [],
        }

    contexte.update({
        'annee_courante': timezone.now().year,
        'notifications': [],
        'nombre_notifications_non_lues': 0,
    })

    return render(request, 'dashboards/caissier_dashboard.html', contexte)

@login_required
def dashboard_restaurant(request):
    """Tableau de bord pour les gestionnaires de restaurant"""
    if not request.user.est_gestionnaire_restaurant:
        messages.warning(request, 'Vous n\'avez pas accès à cette section.')
        return redirect('accounts:dashboard')

    restaurant = request.user.restaurant_gere
    if not restaurant:
        messages.error(request, 'Aucun restaurant assigné à votre compte.')
        return redirect('accounts:dashboard')

    try:
        from apps.transactions.models import JournalConsommation
        from apps.restaurants.models import Menu, Reservation

        aujourdhui = timezone.now().date()
        debut_mois = aujourdhui.replace(day=1)

        # Obtenir les statistiques de consommation
        consommations_jour = JournalConsommation.objects.filter(
            restaurant=restaurant,
            date_consommation__date=aujourdhui
        ).count()

        consommations_mois = JournalConsommation.objects.filter(
            restaurant=restaurant,
            date_consommation__date__gte=debut_mois
        ).count()

        # Obtenir les réservations en attente
        reservations_en_attente = Reservation.objects.filter(
            restaurant=restaurant,
            statut='EN_ATTENTE'
        ).count()

        # Obtenir les menus du jour
        menus_du_jour = Menu.objects.filter(
            restaurant=restaurant,
            date=aujourdhui
        )

        # Obtenir les consommations récentes
        consommations_recentes = JournalConsommation.objects.filter(
            restaurant=restaurant
        ).select_related('client', 'ticket', 'menu_consomme').order_by('-date_consommation')[:10]

        contexte = {
            'consommations_jour': consommations_jour,
            'consommations_mois': consommations_mois,
            'reservations_en_attente': reservations_en_attente,
            'nombre_agences': 0,
            'menus_du_jour': menus_du_jour,
            'consommations_recentes': consommations_recentes,
        }
    except ImportError:
        contexte = {
            'consommations_jour': 0,
            'consommations_mois': 0,
            'reservations_en_attente': 0,
            'nombre_agences': 0,
            'menus_du_jour': [],
            'consommations_recentes': [],
        }

    contexte.update({
        'annee_courante': timezone.now().year,
        'notifications': [],
        'nombre_notifications_non_lues': 0,
    })

    return render(request, 'dashboards/restaurant_dashboard.html', contexte)

@login_required
def dashboard_admin(request):
    """Tableau de bord pour les super administrateurs"""
    if not request.user.est_super_admin:
        messages.warning(request, 'Vous n\'avez pas accès à cette section.')
        return redirect('accounts:dashboard')

    try:
        from apps.transactions.models import TransactionTicket
        from apps.restaurants.models import Restaurant
        from apps.settings.models import JournalAudit

        aujourdhui = timezone.now().date()
        debut_mois = aujourdhui.replace(day=1)

        # Obtenir les totaux
        total_employes = Utilisateur.objects.filter(type_utilisateur='CLIENT', est_actif=True).count()
        nouveaux_employes_mois = Utilisateur.objects.filter(
            type_utilisateur='CLIENT',
            date_inscription__date__gte=debut_mois
        ).count()

        total_directions = Direction.objects.count()
        directions_actives = Direction.objects.filter(est_active=True).count()

        total_agences = Agence.objects.count()
        agences_actives = Agence.objects.filter(est_active=True).count()

        total_restaurants = Restaurant.objects.count()
        restaurants_actifs = Restaurant.objects.filter(statut='ACTIF').count()

        # Obtenir les statistiques des tickets
        transactions_mois = TransactionTicket.objects.filter(
            date_transaction__date__gte=debut_mois,
            statut='TERMINE'
        )
        tickets_vendus_mois = sum(t.nombre_tickets for t in transactions_mois)
        revenu_mois = sum(t.montant_total for t in transactions_mois)

        # Obtenir les tickets consommés
        from apps.tickets.models import Ticket
        tickets_consommes_mois = Ticket.objects.filter(
            statut='CONSOMME',
            date_consommation__date__gte=debut_mois
        ).count()

        # Obtenir les directions principales
        directions_principales = Direction.objects.filter(est_active=True).annotate(
            nombre_employes=Count('employes', filter=Q(employes__type_utilisateur='CLIENT', employes__est_actif=True))
        ).order_by('-nombre_employes')[:5]

        max_nombre = directions_principales.first().nombre_employes if directions_principales else 1
        for direction in directions_principales:
            direction.pourcentage = (direction.nombre_employes / max_nombre * 100) if max_nombre > 0 else 0

        # Obtenir les transactions récentes
        transactions_recentes = TransactionTicket.objects.select_related('client').order_by('-date_transaction')[:10]

        # Obtenir les activités récentes
        try:
            activites_recentes = JournalAudit.objects.select_related('utilisateur').order_by('-cree_le')[:5]
            for activite in activites_recentes:
                if activite.action == 'CREER':
                    activite.icone = 'plus'
                elif activite.action == 'MODIFIER':
                    activite.icone = 'edit'
                elif activite.action == 'SUPPRIMER':
                    activite.icone = 'trash'
                else:
                    activite.icone = 'info-circle'
        except:
            activites_recentes = []

        contexte = {
            'total_employes': total_employes,
            'nouveaux_employes_mois': nouveaux_employes_mois,
            'total_directions': total_directions,
            'directions_actives': directions_actives,
            'total_agences': total_agences,
            'agences_actives': agences_actives,
            'total_restaurants': total_restaurants,
            'restaurants_actifs': restaurants_actifs,
            'tickets_vendus_mois': tickets_vendus_mois,
            'tickets_consommes_mois': tickets_consommes_mois,
            'revenu_mois': revenu_mois,
            'directions_principales': directions_principales,
            'transactions_recentes': transactions_recentes,
            'activites_recentes': activites_recentes,
        }
    except ImportError:
        contexte = {
            'total_employes': Utilisateur.objects.filter(type_utilisateur='CLIENT', est_actif=True).count(),
            'nouveaux_employes_mois': 0,
            'total_directions': Direction.objects.count(),
            'directions_actives': Direction.objects.filter(est_active=True).count(),
            'total_agences': Agence.objects.count(),
            'agences_actives': Agence.objects.filter(est_active=True).count(),
            'total_restaurants': 0,
            'restaurants_actifs': 0,
            'tickets_vendus_mois': 0,
            'tickets_consommes_mois': 0,
            'revenu_mois': 0,
            'directions_principales': [],
            'transactions_recentes': [],
            'activites_recentes': [],
        }

    contexte.update({
        'annee_courante': timezone.now().year,
        'notifications': [],
        'nombre_notifications_non_lues': 0,
    })

    return render(request, 'dashboards/admin_dashboard.html', contexte)


# ============================================
# VUES DE GESTION DES UTILISATEURS
# ============================================
def _generer_mot_de_passe():
    """Génère un mot de passe sécurisé de 12 caractères"""
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    mdp = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
        secrets.choice("!@#$%"),
    ]
    mdp += [secrets.choice(alphabet) for _ in range(8)]
    secrets.SystemRandom().shuffle(mdp)
    return ''.join(mdp)

def _generer_username(matricule, email, prenom, nom):
    """
    Stratégie username :
    1. matricule s'il existe
    2. sinon partie gauche de l'email (avant @)
    3. fallback : prenom.nom en minuscules
    """
    if matricule and matricule.strip():
        return matricule.strip()
    if email:
        base = email.split('@')[0]
        # Vérifier unicité
        username = base
        counter = 1
        while Utilisateur.objects.filter(nom_utilisateur=username).exists():
            username = f"{base}{counter}"
            counter += 1
        return username
    # Fallback
    base = f"{prenom.lower()}.{nom.lower()}".replace(' ', '.')
    username = base
    counter = 1
    while Utilisateur.objects.filter(nom_utilisateur=username).exists():
        username = f"{base}{counter}"
        counter += 1
    return username

@login_required
def users_list(request):
    """Liste des utilisateurs avec filtres et pagination"""
    if not request.user.est_super_admin and not request.user.est_caissier:
        messages.warning(request, 'Vous n\'avez pas accès à cette section.')
        return redirect('accounts:dashboard')

    # Récupérer tous les utilisateurs
    users = Utilisateur.objects.all().select_related('direction', 'agence', 'restaurant_gere')

    # Filtres
    type_filter = request.GET.get('type')
    if type_filter:
        users = users.filter(type_utilisateur=type_filter)

    direction_filter = request.GET.get('direction')
    if direction_filter:
        users = users.filter(direction_id=direction_filter)

    agence_filter = request.GET.get('agence')
    if agence_filter:
        users = users.filter(agence_id=agence_filter)

    statut_filter = request.GET.get('statut')
    if statut_filter == 'actif':
        users = users.filter(est_actif=True)
    elif statut_filter == 'inactif':
        users = users.filter(est_actif=False)

    # Recherche
    search = request.GET.get('search')
    if search:
        users = users.filter(
            Q(prenom__icontains=search) |
            Q(nom__icontains=search) |
            Q(email__icontains=search) |
            Q(matricule__icontains=search)
        )

    # Tri
    users = users.order_by('-date_inscription')

    # Pagination
    paginator = Paginator(users, 20)  # 20 utilisateurs par page
    page = request.GET.get('page')

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)

    # Contexte
    context = {
        'users': users,
        'directions': Direction.objects.filter(est_active=True),
        'agences': Agence.objects.filter(est_active=True),
        'is_paginated': True if paginator.num_pages > 1 else False,
        'page_obj': users,
        'annee_courante': timezone.now().year,
    }

    # Ajouter les restaurants si disponibles
    try:
        from apps.restaurants.models import Restaurant
        context['restaurants'] = Restaurant.objects.filter(statut='ACTIF')
    except ImportError:
        context['restaurants'] = []

    return render(request, 'accounts/users_list.html', context)

@login_required
@require_http_methods(["POST"])
def user_create(request):
    """Créer un nouvel utilisateur (AJAX)"""
    if not request.user.est_super_admin:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        email = request.POST.get('email', '').strip()
        type_utilisateur = request.POST.get('type_utilisateur', 'CLIENT')
        matricule = request.POST.get('matricule', '').strip() or None
        prenom = request.POST.get('prenom', '').strip()
        nom = request.POST.get('nom', '').strip()

        # Validation email unique
        if Utilisateur.objects.filter(email=email).exists():
            return JsonResponse({'error': 'Un utilisateur avec cet email existe déjà'}, status=400)

        # Matricule obligatoire pour CLIENT
        if type_utilisateur == 'CLIENT' and not matricule:
            return JsonResponse({'error': 'Le matricule est obligatoire pour un client (employé)'}, status=400)

        # Vérifier unicité du matricule si fourni
        if matricule and Utilisateur.objects.filter(matricule=matricule).exists():
            return JsonResponse({'error': 'Ce matricule est déjà utilisé'}, status=400)

        # Générer username = matricule en priorité
        username = _generer_username(matricule, email, prenom, nom)

        # Vérifier unicité du username
        if Utilisateur.objects.filter(nom_utilisateur=username).exists():
            base = username
            counter = 1
            while Utilisateur.objects.filter(nom_utilisateur=f"{base}{counter}").exists():
                counter += 1
            username = f"{base}{counter}"

        # Générer mot de passe sécurisé
        # Django utilise PBKDF2 SHA256 via set_password — cryptage automatique
        mot_de_passe = _generer_mot_de_passe()

        user = Utilisateur(
            email=email,
            nom_utilisateur=username,
            prenom=prenom,
            nom=nom,
            telephone=request.POST.get('telephone', ''),
            genre=request.POST.get('genre', ''),
            date_naissance=request.POST.get('date_naissance') or None,
            type_utilisateur=type_utilisateur,
            matricule=matricule,
            departement=request.POST.get('departement', ''),
            poste=request.POST.get('poste', ''),
            adresse=request.POST.get('adresse', ''),
            est_actif=True,
            est_verifie=False,
        )
        # set_password crypte le mot de passe (PBKDF2 SHA256 + salt)
        user.set_password(mot_de_passe)

        # Relations
        if request.POST.get('direction'):
            user.direction_id = request.POST.get('direction')
        if request.POST.get('agence'):
            user.agence_id = request.POST.get('agence')
        if request.POST.get('restaurant_gere'):
            user.restaurant_gere_id = request.POST.get('restaurant_gere')
        if request.FILES.get('photo_profil'):
            user.photo_profil = request.FILES.get('photo_profil')

        user.save()
        ProfilUtilisateur.objects.get_or_create(utilisateur=user)

        # Envoyer email avec les identifiants
        from .signals import envoyer_email_avec_mdp
        envoyer_email_avec_mdp(user, mot_de_passe)

        return JsonResponse({
            'success': True,
            'message': f'Utilisateur {user.get_full_name()} créé. Email d\'accès envoyé à {email}.',
            'user_id': user.id
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def user_edit(request, pk):
    """Éditer un utilisateur (GET: données JSON, POST: mise à jour)"""
    if not request.user.est_super_admin:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        user = Utilisateur.objects.get(pk=pk)
    except Utilisateur.DoesNotExist:
        return JsonResponse({'error': 'Utilisateur non trouvé'}, status=404)

    if request.method == 'GET':
        return JsonResponse({
            'id': user.id,
            'prenom': user.prenom,
            'nom': user.nom,
            'email': user.email,
            'telephone': user.telephone,
            'genre': user.genre,
            'date_naissance': user.date_naissance.isoformat() if user.date_naissance else None,
            'type_utilisateur': user.type_utilisateur,
            'matricule': user.matricule,
            'nom_utilisateur': user.nom_utilisateur,
            'departement': user.departement,
            'poste': user.poste,
            'direction_id': user.direction_id,
            'agence_id': user.agence_id,
            'restaurant_gere_id': user.restaurant_gere_id,
            'adresse': user.adresse,
            'est_actif': user.est_actif,
            'est_verifie': user.est_verifie,
        })

    elif request.method == 'POST':
        try:
            user.prenom          = request.POST.get('prenom', user.prenom)
            user.nom             = request.POST.get('nom', user.nom)
            user.telephone       = request.POST.get('telephone', user.telephone)
            user.genre           = request.POST.get('genre', user.genre)
            user.type_utilisateur = request.POST.get('type_utilisateur', user.type_utilisateur)
            user.departement     = request.POST.get('departement', user.departement)
            user.poste           = request.POST.get('poste', user.poste)
            user.adresse         = request.POST.get('adresse', user.adresse)
            user.est_actif       = request.POST.get('est_actif') == '1'
            user.est_verifie     = request.POST.get('est_verifie') == '1'

            # Matricule — met à jour le username si le matricule change
            new_matricule = request.POST.get('matricule', '').strip() or None
            if new_matricule and new_matricule != user.matricule:
                if Utilisateur.objects.filter(matricule=new_matricule).exclude(pk=user.pk).exists():
                    return JsonResponse({'error': 'Ce matricule est déjà utilisé'}, status=400)
                user.matricule = new_matricule
                user.nom_utilisateur = new_matricule  # Synchroniser le username

            date_naissance = request.POST.get('date_naissance')
            if date_naissance:
                user.date_naissance = date_naissance

            # Relations
            direction_id = request.POST.get('direction')
            user.direction_id = direction_id if direction_id else None

            agence_id = request.POST.get('agence')
            user.agence_id = agence_id if agence_id else None

            restaurant_id = request.POST.get('restaurant_gere')
            user.restaurant_gere_id = restaurant_id if restaurant_id else None

            if request.FILES.get('photo_profil'):
                user.photo_profil = request.FILES.get('photo_profil')

            user.save()

            return JsonResponse({'success': True, 'message': f'{user.get_full_name()} mis à jour'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_http_methods(["POST", "DELETE"])
def user_delete(request, pk):
    """Supprimer un utilisateur (AJAX)"""
    if not request.user.est_super_admin:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        user = Utilisateur.objects.get(pk=pk)

        # Ne pas permettre la suppression de soi-même
        if user.id == request.user.id:
            return JsonResponse({
                'error': 'Vous ne pouvez pas supprimer votre propre compte'
            }, status=400)

        user_name = user.get_full_name()
        user.delete()

        messages.success(request, f'Utilisateur {user_name} supprimé avec succès.')
        return JsonResponse({
            'success': True,
            'message': 'Utilisateur supprimé avec succès'
        })

    except Utilisateur.DoesNotExist:
        return JsonResponse({'error': 'Utilisateur non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def user_detail(request, pk):
    """Afficher les détails d'un utilisateur"""
    if not request.user.est_super_admin and request.user.id != pk:
        messages.warning(request, 'Vous n\'avez pas accès à cette page.')
        return redirect('accounts:dashboard')

    user = get_object_or_404(Utilisateur, pk=pk)

    context = {
        'user_detail': user,
        'annee_courante': timezone.now().year,
    }

    return render(request, 'accounts/modals/user_detail.html', context)


# ============================================
# VUES DE PROFIL
# ============================================

@login_required
def profile_view(request):
    """Afficher le profil de l'utilisateur"""
    # Créer le profil s'il n'existe pas
    profil, created = ProfilUtilisateur.objects.get_or_create(utilisateur=request.user)

    context = {
        'annee_courante': timezone.now().year,
        'notifications': [],
        'nombre_notifications_non_lues': 0,
    }

    return render(request, 'accounts/profile.html', context)

@login_required
@require_http_methods(["POST"])
def profile_update(request):
    """Mettre à jour le profil utilisateur (AJAX)"""
    try:
        user = request.user

        # Mise à jour des informations utilisateur
        user.prenom = request.POST.get('prenom', user.prenom)
        user.nom = request.POST.get('nom', user.nom)
        user.telephone = request.POST.get('telephone', user.telephone)
        user.genre = request.POST.get('genre', user.genre)

        # Date de naissance
        date_naissance = request.POST.get('date_naissance')
        if date_naissance:
            user.date_naissance = date_naissance

        user.adresse = request.POST.get('adresse', user.adresse)

        # Photo de profil
        if request.FILES.get('photo_profil'):
            user.photo_profil = request.FILES.get('photo_profil')

        user.save()

        # Mise à jour du profil
        profil, created = ProfilUtilisateur.objects.get_or_create(utilisateur=user)
        profil.notification_email = request.POST.get('notification_email') == '1'
        profil.notification_sms = request.POST.get('notification_sms') == '1'
        profil.contact_urgence_nom = request.POST.get('contact_urgence_nom', '')
        profil.contact_urgence_telephone = request.POST.get('contact_urgence_telephone', '')
        profil.save()

        messages.success(request, 'Profil mis à jour avec succès.')
        return JsonResponse({
            'success': True,
            'message': 'Profil mis à jour avec succès'
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_http_methods(["POST"])
def change_password(request):
    """Changer le mot de passe de l'utilisateur (AJAX)"""
    try:
        data = json.loads(request.body)

        ancien_mdp = data.get('ancien_mot_de_passe')
        nouveau_mdp = data.get('nouveau_mot_de_passe')
        nouveau_mdp2 = data.get('nouveau_mot_de_passe2')

        # Vérifier l'ancien mot de passe
        if not request.user.check_password(ancien_mdp):
            return JsonResponse({
                'error': 'Mot de passe actuel incorrect'
            }, status=400)

        # Vérifier que les nouveaux mots de passe correspondent
        if nouveau_mdp != nouveau_mdp2:
            return JsonResponse({
                'error': 'Les nouveaux mots de passe ne correspondent pas'
            }, status=400)

        # Vérifier la longueur minimale
        if len(nouveau_mdp) < 8:
            return JsonResponse({
                'error': 'Le mot de passe doit contenir au moins 8 caractères'
            }, status=400)

        # Changer le mot de passe
        request.user.set_password(nouveau_mdp)
        request.user.save()

        # Reconnecter l'utilisateur
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, request.user)

        messages.success(request, 'Mot de passe modifié avec succès.')
        return JsonResponse({
            'success': True,
            'message': 'Mot de passe modifié avec succès'
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Données invalides'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ============================================
# VUES DE GESTION DES DIRECTIONS
# ============================================

@login_required
def directions_list(request):
    """Liste des directions avec filtres"""
    if not request.user.est_super_admin:
        messages.warning(request, 'Vous n\'avez pas accès à cette section.')
        return redirect('accounts:dashboard')

    directions = Direction.objects.all().annotate(
        nombre_employes=Count('employes', filter=Q(
            employes__type_utilisateur='CLIENT',
            employes__est_actif=True
        )),
        nombre_agences=Count('agences', filter=Q(agences__est_active=True))
    )

    # Filtres
    statut_filter = request.GET.get('statut')
    if statut_filter == 'actif':
        directions = directions.filter(est_active=True)
    elif statut_filter == 'inactif':
        directions = directions.filter(est_active=False)

    # Recherche
    search = request.GET.get('search')
    if search:
        directions = directions.filter(
            Q(nom__icontains=search) |
            Q(code__icontains=search)
        )

    directions = directions.order_by('nom')

    context = {
        'directions': directions,
        'annee_courante': timezone.now().year,
        'utilisateurs': Utilisateur.objects.filter(
            type_utilisateur='SUPER_ADMIN',
            est_actif=True
        ),
    }

    return render(request, 'accounts/directions_list.html', context)

@login_required
@require_http_methods(["POST"])
def direction_create(request):
    """Créer une nouvelle direction (AJAX)"""
    if not request.user.est_super_admin:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        # Vérifier que le code n'existe pas déjà
        if Direction.objects.filter(code=request.POST.get('code')).exists():
            return JsonResponse({
                'error': 'Une direction avec ce code existe déjà'
            }, status=400)

        direction = Direction.objects.create(
            nom=request.POST.get('nom'),
            code=request.POST.get('code'),
            description=request.POST.get('description', ''),
            telephone=request.POST.get('telephone', ''),
            email=request.POST.get('email', ''),
        )

        if request.POST.get('directeur'):
            direction.directeur_id = request.POST.get('directeur')
            direction.save()

        messages.success(request, f'Direction {direction.nom} créée avec succès.')
        return JsonResponse({
            'success': True,
            'message': 'Direction créée avec succès',
            'direction_id': direction.id
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def direction_edit(request, pk):
    """Éditer une direction (GET: données, POST: mise à jour)"""
    if not request.user.est_super_admin:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        direction = Direction.objects.get(pk=pk)
    except Direction.DoesNotExist:
        return JsonResponse({'error': 'Direction non trouvée'}, status=404)

    if request.method == 'GET':
        data = {
            'id': direction.id,
            'nom': direction.nom,
            'code': direction.code,
            'description': direction.description,
            'telephone': direction.telephone,
            'email': direction.email,
            'directeur_id': direction.directeur_id,
            'est_active': direction.est_active,
        }
        return JsonResponse(data)

    elif request.method == 'POST':
        try:
            direction.nom = request.POST.get('nom', direction.nom)
            direction.description = request.POST.get('description', direction.description)
            direction.telephone = request.POST.get('telephone', direction.telephone)
            direction.email = request.POST.get('email', direction.email)
            direction.est_active = request.POST.get('est_active') == '1'

            directeur_id = request.POST.get('directeur')
            if directeur_id:
                direction.directeur_id = directeur_id
            else:
                direction.directeur = None

            direction.save()

            messages.success(request, f'Direction {direction.nom} mise à jour avec succès.')
            return JsonResponse({
                'success': True,
                'message': 'Direction mise à jour avec succès'
            })

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_http_methods(["POST", "DELETE"])
def direction_delete(request, pk):
    """Supprimer une direction (AJAX)"""
    if not request.user.est_super_admin:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        direction = Direction.objects.get(pk=pk)

        # Vérifier s'il y a des employés ou agences liés
        if direction.employes.exists():
            return JsonResponse({
                'error': 'Impossible de supprimer une direction qui a des employés'
            }, status=400)

        if direction.agences.exists():
            return JsonResponse({
                'error': 'Impossible de supprimer une direction qui a des agences'
            }, status=400)

        direction_name = direction.nom
        direction.delete()

        messages.success(request, f'Direction {direction_name} supprimée avec succès.')
        return JsonResponse({
            'success': True,
            'message': 'Direction supprimée avec succès'
        })

    except Direction.DoesNotExist:
        return JsonResponse({'error': 'Direction non trouvée'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def direction_detail(request, pk):
    """Afficher les détails d'une direction"""
    if not request.user.est_super_admin:
        messages.warning(request, 'Vous n\'avez pas accès à cette page.')
        return redirect('accounts:dashboard')

    direction = get_object_or_404(Direction, pk=pk)

    context = {
        'direction': direction,
        'annee_courante': timezone.now().year,
    }

    return render(request, 'accounts/direction_detail.html', context)


# ============================================
# VUES DE GESTION DES AGENCES
# ============================================

@login_required
def agences_list(request):
    """Liste des agences avec filtres"""
    if not request.user.est_super_admin:
        messages.warning(request, 'Vous n\'avez pas accès à cette section.')
        return redirect('accounts:dashboard')

    agences = Agence.objects.all().select_related(
        'direction',
        'responsable',
        'agence_parente'
    ).annotate(
        nombre_employes=Count('employes', filter=Q(
            employes__type_utilisateur='CLIENT',
            employes__est_actif=True
        ))
    )

    # Filtres
    type_filter = request.GET.get('type')
    if type_filter:
        agences = agences.filter(type_agence=type_filter)

    direction_filter = request.GET.get('direction')
    if direction_filter:
        agences = agences.filter(direction_id=direction_filter)

    statut_filter = request.GET.get('statut')
    if statut_filter == 'actif':
        agences = agences.filter(est_active=True)
    elif statut_filter == 'inactif':
        agences = agences.filter(est_active=False)

    # Recherche
    search = request.GET.get('search')
    if search:
        agences = agences.filter(
            Q(nom__icontains=search) |
            Q(code__icontains=search) |
            Q(ville__icontains=search)
        )

    agences = agences.order_by('nom')

    context = {
        'agences': agences,
        'directions': Direction.objects.filter(est_active=True),
        'annee_courante': timezone.now().year,
        'responsables': Utilisateur.objects.filter(
            type_utilisateur__in=['SUPER_ADMIN', 'CAISSIER'],
            est_actif=True
        ),
    }

    return render(request, 'accounts/agences_list.html', context)

@login_required
@require_http_methods(["POST"])
def agence_create(request):
    """Créer une nouvelle agence (AJAX)"""
    if not request.user.est_super_admin:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        code = request.POST.get('code', '').strip()
        if not code:
            return JsonResponse({'error': 'Le code est obligatoire'}, status=400)

        if Agence.objects.filter(code=code).exists():
            return JsonResponse({'error': 'Une agence avec ce code existe déjà'}, status=400)

        agence = Agence.objects.create(
            nom=request.POST.get('nom', '').strip(),
            code=code,
            type_agence=request.POST.get('type_agence', 'LOCALE'),
            adresse=request.POST.get('adresse', '').strip(),
            ville=request.POST.get('ville', '').strip(),
            region=request.POST.get('region', '') or '',
            telephone=request.POST.get('telephone', '').strip(),
            email=request.POST.get('email', '') or '',
        )

        if request.POST.get('direction'):
            agence.direction_id = request.POST.get('direction')
        if request.POST.get('responsable'):
            agence.responsable_id = request.POST.get('responsable')

        agence.save()

        return JsonResponse({
            'success': True,
            'message': f'Agence {agence.nom} créée avec succès',
            'agence_id': agence.id
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def agence_edit(request, pk):
    """Éditer une agence (GET: données, POST: mise à jour)"""
    if not request.user.est_super_admin:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        agence = Agence.objects.get(pk=pk)
    except Agence.DoesNotExist:
        return JsonResponse({'error': 'Agence non trouvée'}, status=404)

    if request.method == 'GET':
        data = {
            'id': agence.id,
            'nom': agence.nom,
            'code': agence.code,
            'type_agence': agence.type_agence,
            'adresse': agence.adresse,
            'ville': agence.ville,
            'region': agence.region,
            'telephone': agence.telephone,
            'email': agence.email,
            'direction_id': agence.direction_id,
            'agence_parente_id': agence.agence_parente_id,
            'responsable_id': agence.responsable_id,
            'est_active': agence.est_active,
        }
        return JsonResponse(data)

    elif request.method == 'POST':
        try:
            agence.nom = request.POST.get('nom', agence.nom)
            agence.type_agence = request.POST.get('type_agence', agence.type_agence)
            agence.adresse = request.POST.get('adresse', agence.adresse)
            agence.ville = request.POST.get('ville', agence.ville)
            agence.region = request.POST.get('region', agence.region)
            agence.telephone = request.POST.get('telephone', agence.telephone)
            agence.email = request.POST.get('email', agence.email)
            agence.est_active = request.POST.get('est_active') == '1'

            # Relations
            direction_id = request.POST.get('direction')
            if direction_id:
                agence.direction_id = direction_id
            else:
                agence.direction = None

            agence_parente_id = request.POST.get('agence_parente')
            if agence_parente_id:
                agence.agence_parente_id = agence_parente_id
            else:
                agence.agence_parente = None

            responsable_id = request.POST.get('responsable')
            if responsable_id:
                agence.responsable_id = responsable_id
            else:
                agence.responsable = None

            agence.save()

            messages.success(request, f'Agence {agence.nom} mise à jour avec succès.')
            return JsonResponse({
                'success': True,
                'message': 'Agence mise à jour avec succès'
            })

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_http_methods(["POST", "DELETE"])
def agence_delete(request, pk):
    """Supprimer une agence (AJAX)"""
    if not request.user.est_super_admin:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        agence = Agence.objects.get(pk=pk)

        # Vérifier s'il y a des employés liés
        if agence.employes.exists():
            return JsonResponse({
                'error': 'Impossible de supprimer une agence qui a des employés'
            }, status=400)

        # Vérifier s'il y a des sous-agences
        if agence.sous_agences.exists():
            return JsonResponse({
                'error': 'Impossible de supprimer une agence qui a des sous-agences'
            }, status=400)

        agence_name = agence.nom
        agence.delete()

        messages.success(request, f'Agence {agence_name} supprimée avec succès.')
        return JsonResponse({
            'success': True,
            'message': 'Agence supprimée avec succès'
        })

    except Agence.DoesNotExist:
        return JsonResponse({'error': 'Agence non trouvée'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def agence_detail(request, pk):
    """Afficher les détails d'une agence"""
    if not request.user.est_super_admin:
        messages.warning(request, 'Vous n\'avez pas accès à cette page.')
        return redirect('accounts:dashboard')

    agence = get_object_or_404(Agence, pk=pk)

    context = {
        'agence': agence,
        'annee_courante': timezone.now().year,
    }

    return render(request, 'accounts/agence_detail.html', context)


# ============================================
# VUES D'EXPORT
# ============================================

@login_required
def export_users_pdf(request):
    """Exporter la liste des utilisateurs en PDF"""
    if not request.user.est_super_admin:
        messages.warning(request, 'Vous n\'avez pas accès à cette fonction.')
        return redirect('accounts:users_list')

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    # Créer le buffer
    buffer = BytesIO()

    # Créer le document PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a472a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    # Titre
    title = Paragraph("Liste des Utilisateurs", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Récupérer les utilisateurs
    users = Utilisateur.objects.all().select_related('direction', 'agence')

    # Données du tableau
    data = [['#', 'Nom', 'Email', 'Type', 'Direction', 'Statut']]

    for idx, user in enumerate(users, 1):
        data.append([
            str(idx),
            user.get_full_name(),
            user.email,
            user.get_type_utilisateur_display(),
            user.direction.nom if user.direction else '-',
            'Actif' if user.est_actif else 'Inactif'
        ])

    # Créer le tableau
    table = Table(data, colWidths=[0.5 * inch, 1.5 * inch, 2 * inch, 1.5 * inch, 1.5 * inch, 1 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a472a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)

    # Construire le PDF
    doc.build(elements)

    # Retourner la réponse
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="utilisateurs_{timezone.now().strftime("%Y%m%d")}.pdf"'

    return response

@login_required
def export_users_excel(request):
    """Exporter la liste des utilisateurs en Excel"""
    if not request.user.est_super_admin:
        messages.warning(request, 'Vous n\'avez pas accès à cette fonction.')
        return redirect('accounts:users_list')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"

    # Style pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a472a", end_color="1a472a", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes
    headers = ['#', 'Prénom', 'Nom', 'Email', 'Téléphone', 'Type', 'Matricule',
               'Direction', 'Agence', 'Statut', 'Date d\'inscription']
    ws.append(headers)

    # Appliquer le style aux en-têtes
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Récupérer les utilisateurs
    users = Utilisateur.objects.all().select_related('direction', 'agence')

    # Ajouter les données
    for idx, user in enumerate(users, 1):
        ws.append([
            idx,
            user.prenom,
            user.nom,
            user.email,
            user.telephone,
            user.get_type_utilisateur_display(),
            user.matricule or '-',
            user.direction.nom if user.direction else '-',
            user.agence.nom if user.agence else '-',
            'Actif' if user.est_actif else 'Inactif',
            user.date_inscription.strftime('%d/%m/%Y') if user.date_inscription else '-'
        ])

    # Ajuster la largeur des colonnes
    column_widths = [5, 15, 15, 25, 15, 20, 15, 20, 20, 10, 15]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Retourner la réponse
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="utilisateurs_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    return response

@login_required
def export_directions_pdf(request):
    """Exporter la liste des directions en PDF"""
    if not request.user.est_super_admin:
        messages.warning(request, 'Vous n\'avez pas accès à cette fonction.')
        return redirect('accounts:directions_list')

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a472a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    title = Paragraph("Liste des Directions", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))

    directions = Direction.objects.all()

    data = [['#', 'Nom', 'Code', 'Directeur', 'Employés', 'Statut']]

    for idx, direction in enumerate(directions, 1):
        data.append([
            str(idx),
            direction.nom,
            direction.code,
            direction.directeur.get_full_name() if direction.directeur else '-',
            str(direction.total_employes),
            'Active' if direction.est_active else 'Inactive'
        ])

    table = Table(data, colWidths=[0.5 * inch, 2 * inch, 1 * inch, 1.5 * inch, 1 * inch, 1 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a472a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="directions_{timezone.now().strftime("%Y%m%d")}.pdf"'

    return response

@login_required
def export_agencies_excel(request):
    """Exporter la liste des agences en Excel"""
    if not request.user.est_super_admin:
        messages.warning(request, 'Vous n\'avez pas accès à cette fonction.')
        return redirect('accounts:agencies_list')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Agences"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a472a", end_color="1a472a", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ['#', 'Nom', 'Code', 'Type', 'Ville', 'Direction', 'Responsable', 'Employés', 'Statut']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    agences = Agence.objects.all().select_related('direction', 'responsable')

    for idx, agence in enumerate(agences, 1):
        ws.append([
            idx,
            agence.nom,
            agence.code,
            agence.get_type_agence_display(),
            agence.ville,
            agence.direction.nom if agence.direction else '-',
            agence.responsable.get_full_name() if agence.responsable else '-',
            agence.total_employes,
            'Active' if agence.est_active else 'Inactive'
        ])

    column_widths = [5, 20, 10, 15, 15, 20, 20, 10, 10]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="agences_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    return response


# ============================================
# API ViewSets
# ============================================

class UtilisateurViewSet(viewsets.ModelViewSet):
    """ViewSet pour le modèle Utilisateur"""
    queryset = Utilisateur.objects.all()
    serializer_class = UtilisateurSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'create':
            return UtilisateurCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UtilisateurMiseAJourSerializer
        return UtilisateurSerializer

    @action(detail=False, methods=['get'])
    def moi(self, request):
        """Obtenir l'utilisateur actuel"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def changer_mot_de_passe(self, request, pk=None):
        """Changer le mot de passe de l'utilisateur"""
        utilisateur = self.get_object()
        serializer = ChangerMotDePasseSerializer(data=request.data, context={'request': request})

        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Mot de passe modifié avec succès.'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def employes_actifs(self, request):
        """Obtenir les employés actifs"""
        employes = Utilisateur.objects.filter(type_utilisateur='CLIENT', est_actif=True)
        page = self.paginate_queryset(employes)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(employes, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def par_agence(self, request):
        """Obtenir les utilisateurs par agence"""
        agence_id = request.query_params.get('agence_id')
        if agence_id:
            utilisateurs = Utilisateur.objects.filter(agence_id=agence_id)
        else:
            utilisateurs = Utilisateur.objects.all()

        page = self.paginate_queryset(utilisateurs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(utilisateurs, many=True)
        return Response(serializer.data)

class DirectionViewSet(viewsets.ModelViewSet):
    """ViewSet pour le modèle Direction"""
    queryset = Direction.objects.all()
    serializer_class = DirectionSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['get'])
    def employes(self, request, pk=None):
        """Obtenir les employés d'une direction"""
        direction = self.get_object()
        employes = direction.employes.filter(type_utilisateur='CLIENT')
        page = self.paginate_queryset(employes)
        if page is not None:
            serializer = UtilisateurSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = UtilisateurSerializer(employes, many=True)
        return Response(serializer.data)

class AgenceViewSet(viewsets.ModelViewSet):
    """ViewSet pour le modèle Agence"""
    queryset = Agence.objects.all()
    serializer_class = AgenceSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['get'])
    def employes(self, request, pk=None):
        """Obtenir les employés d'une agence"""
        agence = self.get_object()
        employes = agence.employes.filter(type_utilisateur='CLIENT')
        page = self.paginate_queryset(employes)
        if page is not None:
            serializer = UtilisateurSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = UtilisateurSerializer(employes, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def avec_capacite(self, request):
        """Obtenir les agences avec capacité disponible"""
        agences = Agence.objects.filter(est_active=True)
        agences_avec_capacite = [agence for agence in agences if agence.a_capacite]

        page = self.paginate_queryset(agences_avec_capacite)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(agences_avec_capacite, many=True)
        return Response(serializer.data)

class ProfilUtilisateurViewSet(viewsets.ModelViewSet):
    """ViewSet pour le modèle ProfilUtilisateur"""
    queryset = ProfilUtilisateur.objects.all()
    serializer_class = ProfilUtilisateurSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ['update', 'partial_update']:
            return ProfilUtilisateurMiseAJourSerializer
        return ProfilUtilisateurSerializer

    @action(detail=False, methods=['get'])
    def mon_profil(self, request):
        """Obtenir le profil de l'utilisateur actuel"""
        profil, cree = ProfilUtilisateur.objects.get_or_create(utilisateur=request.user)
        serializer = self.get_serializer(profil)
        return Response(serializer.data)


# ============================================
# Vues API supplémentaires
# ============================================

@api_view(['POST'])
@permission_classes([AllowAny])
def api_connexion(request):
    """API pour la connexion (retourne JWT tokens)"""
    serializer = LoginSerializer(data=request.data, context={'request': request})

    if serializer.is_valid():
        utilisateur = serializer.validated_data['utilisateur']

        # Générer les tokens JWT
        refresh = RefreshToken.for_user(utilisateur)

        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'utilisateur': UtilisateurSerializer(utilisateur).data
        })

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_deconnexion(request):
    """API pour la déconnexion"""
    try:
        refresh_token = request.data.get('refresh')
        token = RefreshToken(refresh_token)
        token.blacklist()
        return Response({'message': 'Déconnexion réussie.'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def api_mise_a_jour_profil(request):
    """API pour mettre à jour le profil utilisateur"""
    utilisateur_serializer = UtilisateurMiseAJourSerializer(
        request.user,
        data=request.data.get('utilisateur', {}),
        partial=True
    )

    profil_serializer = ProfilUtilisateurMiseAJourSerializer(
        request.user.profil,
        data=request.data.get('profil', {}),
        partial=True
    )

    if utilisateur_serializer.is_valid() and profil_serializer.is_valid():
        utilisateur_serializer.save()
        profil_serializer.save()

        return Response({
            'message': 'Profil mis à jour avec succès.',
            'utilisateur': UtilisateurSerializer(request.user).data,
            'profil': ProfilUtilisateurSerializer(request.user.profil).data
        })

    errors = {}
    if utilisateur_serializer.errors:
        errors['utilisateur'] = utilisateur_serializer.errors
    if profil_serializer.errors:
        errors['profil'] = profil_serializer.errors

    return Response(errors, status=status.HTTP_400_BAD_REQUEST)